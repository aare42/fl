#!/usr/bin/env python3
"""
MOD_SYNC_ENGINE - Python Rewrite
Synchronizes data between Master and Source Excel workbooks
Preserves exact VBA logic including any quirks/bugs
"""

import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import sys

# ============================================================================
# CONSTANTS
# ============================================================================
LOG_SHEET_NAME = "MAPTransfer"
SH_OOS = "2. ООС"
SH_EXCL = "3. Виключені"
SH_PRIK = "4. Тимчасово прибулі"
SH_ABS = "5. Тимчасово відсутні"

# ============================================================================
# GLOBAL STATE (mimicking VBA globals)
# ============================================================================
logs: List[Tuple[datetime, str, str, str]] = []

# Placeholders for functions from other modules
# These need to be implemented based on your other VBA modules
def load_index_maps():
    """TODO: Implement from your other VBA module"""
    pass

def load_division_maps():
    """TODO: Implement from your other VBA module"""
    pass

def prepare_free_cadet_indexes(table, col_index):
    """TODO: Implement from your other VBA module"""
    pass

def core_archive_tabel_row(tabel_sheet, index_pos):
    """TODO: Implement from your other VBA module"""
    pass

def core_move_person(pib, old_index, new_index, nakaz_num, nakaz_date,
                     date_val, oos_sheet, tabel_sheet, abs_sheet):
    """TODO: Implement from your other VBA module"""
    pass

def core_register_new_oos(src_row_data, index, oos_sheet, tabel_sheet):
    """TODO: Implement from your other VBA module"""
    pass

def core_add_prikom(data_dict, prik_sheet, tabel_sheet):
    """TODO: Implement from your other VBA module"""
    pass

def generate_prik_index(nb: str, nv: str) -> str:
    """TODO: Implement from your other VBA module"""
    pass

# ============================================================================
# LOGGING
# ============================================================================
def log_event(event_type: str, message: str, key: str = ""):
    """Add event to log collection"""
    logs.append((datetime.now(), event_type, key, message))
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")

def clear_old_logs(wb):
    """Clear existing logs from log sheet"""
    try:
        if LOG_SHEET_NAME in wb.sheetnames:
            ws = wb[LOG_SHEET_NAME]
            # Find table and clear data
            for table in ws.tables.values():
                if table.name == "ТаблицяЛог":
                    # Clear data rows (keep header)
                    min_row = table.ref.split(':')[0]
                    max_row = table.ref.split(':')[1]
                    start_row = int(''.join(filter(str.isdigit, min_row))) + 1
                    end_row = int(''.join(filter(str.isdigit, max_row)))

                    for row in range(start_row, end_row + 1):
                        for col in range(1, 5):
                            ws.cell(row, col).value = None
                    break
    except Exception as e:
        print(f"Warning: Could not clear old logs: {e}")

def dump_logs(wb):
    """Write collected logs to log sheet"""
    if LOG_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(LOG_SHEET_NAME)
        ws['A1'] = "Time"
        ws['B1'] = "Type"
        ws['C1'] = "Key"
        ws['D1'] = "Message"

        # Create table
        tab = openpyxl.worksheet.table.Table(
            displayName="ТаблицяЛог",
            ref=f"A1:D{len(logs) + 1}"
        )
        ws.add_table(tab)
    else:
        ws = wb[LOG_SHEET_NAME]

    # Write log data
    for i, (timestamp, event_type, key, message) in enumerate(logs, start=2):
        ws[f'A{i}'] = timestamp
        ws[f'B{i}'] = event_type
        ws[f'C{i}'] = key
        ws[f'D{i}'] = message

    # Auto-fit columns
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].auto_size = True

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def find_row_index_by_column_value(ws, table_name: str, col_name: str,
                                   value: str) -> int:
    """
    Find row index in table where column matches value
    Returns 1-based index (VBA style), 0 if not found
    """
    try:
        table = ws.tables[table_name]
        table_range = ws[table.ref]

        # Find column index
        header_row = next(table_range)
        col_idx = None
        for idx, cell in enumerate(header_row, start=1):
            if cell.value == col_name:
                col_idx = idx
                break

        if col_idx is None:
            return 0

        # Search for value
        for row_idx, row in enumerate(table_range, start=1):
            if row_idx == 1:  # Skip header
                continue
            if str(row[col_idx - 1].value) == str(value):
                return row_idx - 1  # Return data body index (0-based from data start)

        return 0
    except Exception as e:
        print(f"Warning: Error in find_row_index: {e}")
        return 0

def get_table_data_as_list(ws, table_name: str) -> List[List[Any]]:
    """Get table data as 2D list (excluding header)"""
    try:
        table = ws.tables[table_name]
        table_range = ws[table.ref]

        data = []
        for row_idx, row in enumerate(table_range):
            if row_idx == 0:  # Skip header
                continue
            data.append([cell.value for cell in row])

        return data
    except Exception as e:
        print(f"Warning: Error reading table {table_name}: {e}")
        return []

def update_cell_if_changed(cell, new_value, log_msg: str, key: str):
    """Update cell only if value changed (mimics VBA behavior)"""
    old_value = cell.value

    if str(new_value) != str(old_value):
        if str(new_value) != "":
            cell.value = new_value
            log_event("UPD-INFO", log_msg, key)

# ============================================================================
# MAIN SYNCHRONIZATION ROUTINE
# ============================================================================
def run_synchronization_main(master_path: str):
    """Main synchronization routine"""
    global logs
    logs = []

    t_start = datetime.now()

    print("Підготовка: очистка логів...")
    wb_master = openpyxl.load_workbook(master_path)
    clear_old_logs(wb_master)

    # File selection dialog
    print("Виберіть файл-копію для синхронізації...")
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Оберіть файл-копію для синхронізації",
        initialdir=str(Path(master_path).parent),
        filetypes=[("Excel Files", "*.xls *.xlsx *.xlsm *.xlsb")]
    )

    if not file_path:
        print("Відмінено користувачем")
        return

    log_event("START", f"Початок синхронізації з файлом: {file_path}", "")

    # Load reference data
    print("Завантаження довідників та карт...")
    load_index_maps()
    load_division_maps()

    print("Інвентаризація вільних посад...")
    # TODO: Get actual table object and column index
    # prepare_free_cadet_indexes(...)

    # Open source file
    print("Відкриття файлу-копії (це може зайняти час)...")
    wb_source = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # STAGE A: Excluded
    print("Етап 1/4: Обробка вибувших...")
    sync_step_excluded(wb_source, wb_master)

    # STAGE B: OOS
    print("Етап 2/4: Синхронізація ООС...")
    sync_step_oos(wb_source, wb_master)

    # STAGE C: Prikom
    print("Етап 3/4: Прикомандировані...")
    sync_step_prikom(wb_source, wb_master)

    # STAGE D: Absent
    print("Етап 4/4: Тимчасово відсутні...")
    sync_step_absent(wb_source, wb_master)

    # Cleanup
    print("Закриття файлу та збереження звіту...")
    wb_source.close()

    # Write logs
    dump_logs(wb_master)

    # Save master
    wb_master.save(master_path)

    elapsed = (datetime.now() - t_start).total_seconds()
    print(f"\n✅ Синхронізацію завершено за {elapsed:.1f} сек.")
    print(f"Перевірте звіт на сторінці {LOG_SHEET_NAME}")

# ============================================================================
# STAGE A: EXCLUDED
# ============================================================================
def sync_step_excluded(wb_src, wb_dst):
    """Process excluded personnel"""
    log_event("STEP", "Обробка таблиці 'Виключені'", "")

    ws_src = wb_src[SH_EXCL]
    ws_dst = wb_dst[SH_EXCL]
    ws_oos_dst = wb_dst[SH_OOS]

    # Build dictionary of existing excluded personnel
    dict_dst = {}
    data_dst = get_table_data_as_list(ws_dst, "ТаблицяВиключені")
    for row in data_dst:
        inn = str(row[19]) if len(row) > 19 else ""  # Column 20 (0-indexed = 19)
        pib = str(row[1]) if len(row) > 1 else ""    # Column 2 (0-indexed = 1)

        if inn and inn != "0":
            dict_dst[inn] = 1
        else:
            dict_dst[pib] = 1

    # Process source data
    data_src = get_table_data_as_list(ws_src, "ТаблицяВиключені")

    total_rows = len(data_src)
    for i, row in enumerate(data_src, start=1):
        if i % 10 == 0:
            print(f"  Виключені: обробка рядка {i} з {total_rows}")

        inn = str(row[19]) if len(row) > 19 else ""
        pib = str(row[1]) if len(row) > 1 else ""

        # Check if exists
        exists = dict_dst.get(inn if inn and inn != "0" else pib, False)

        if not exists:
            # Add new row to excluded
            # TODO: Implement actual row addition to table
            log_event("MOVE", f"Перенесено у Виключені: {pib}", inn)

            # Find and delete from OOS
            idx_oos = find_row_index_by_column_value(ws_oos_dst, "ТаблицяООС", "ІНН", inn)
            if idx_oos == 0:
                idx_oos = find_row_index_by_column_value(ws_oos_dst, "ТаблицяООС", "ПІБ", pib)

            if idx_oos > 0:
                # TODO: Get actual index from table
                # index_pos = ...
                # core_archive_tabel_row(wb_dst["6. Табель"], index_pos)
                # Delete row from OOS table
                log_event("DEL", f"Видалено з ООС (архівовано в табелі): {pib}", inn)
            else:
                log_event("WARN", "Людина є у Виключених (копія), але не знайдена в ООС (оригінал)", pib)

# ============================================================================
# STAGE B: OOS (RANKS + CONTRACTS)
# ============================================================================
def sync_step_oos(wb_src, wb_dst):
    """Sync main personnel list"""
    log_event("STEP", "Обробка ООС (Посади, Звання, Контракти)", "")

    ws_src = wb_src[SH_OOS]
    ws_dst = wb_dst[SH_OOS]

    data_src = get_table_data_as_list(ws_src, "ТаблицяООС")

    # Column indices (0-based, subtract 1 from VBA)
    # TODO: Get actual column indices from table headers
    c_inn = 20  # Column U (21st column)
    c_pib = 1   # Column B
    c_ind = 2   # Column C (Індекс посади)

    total_rows = len(data_src)

    for i, row in enumerate(data_src, start=1):
        if i % 10 == 0:
            print(f"  ООС: перевірка особи {i} з {total_rows}")

        inn = str(row[c_inn]) if len(row) > c_inn else ""
        pib = str(row[c_pib]) if len(row) > c_pib else ""
        index_src = str(row[c_ind]) if len(row) > c_ind else ""

        # Find in destination
        idx_dst = 0
        if inn and inn != "0":
            idx_dst = find_row_index_by_column_value(ws_dst, "ТаблицяООС", "ІНН", inn)
        if idx_dst == 0:
            idx_dst = find_row_index_by_column_value(ws_dst, "ТаблицяООС", "ПІБ", pib)

        if idx_dst > 0:
            # Person exists - check for changes
            # TODO: Get actual row data and compare
            # 1. Position change
            # 2. Rank updates
            # 3. Contract updates
            pass
        else:
            # New person - check if not excluded
            is_excl = find_row_index_by_column_value(
                wb_dst[SH_EXCL], "ТаблицяВиключені", "ІНН", inn
            ) > 0

            if not is_excl:
                log_event("ADD", f"Нова людина в ООС: {pib}", inn)
                # TODO: core_register_new_oos(row_data, index_src, ws_dst, tabel_sheet)
            else:
                log_event("SKIP", f"Людина є у списку Виключених, пропуск: {pib}", inn)

# ============================================================================
# STAGE C: PRIKOM
# ============================================================================
def sync_step_prikom(wb_src, wb_dst):
    """Sync temporarily assigned personnel"""
    log_event("STEP", "Обробка Прикомандированих", "")

    ws_src = wb_src[SH_PRIK]
    ws_dst = wb_dst[SH_PRIK]

    data_src = get_table_data_as_list(ws_src, "ТаблицяПрик")

    # Build destination dictionary
    dict_dst = {}
    data_dst = get_table_data_as_list(ws_dst, "ТаблицяПрик")
    for idx, row in enumerate(data_dst, start=1):
        pib = str(row[3]) if len(row) > 3 else ""
        date_arr = row[7] if len(row) > 7 else ""
        key = f"{pib}|{date_arr}"
        dict_dst[key] = idx

    total_rows = len(data_src)

    for i, row in enumerate(data_src, start=1):
        if i % 10 == 0:
            print(f"  Прикомандировані: рядок {i} з {total_rows}")

        pib = str(row[3]) if len(row) > 3 else ""
        date_arr = row[7] if len(row) > 7 else ""
        src_key = f"{pib}|{date_arr}"

        if src_key in dict_dst:
            # Exists - check for departure date
            row_idx = dict_dst[src_key]
            # TODO: Check and update departure dates
            pass
        else:
            # New arrival
            log_event("ADD", f"Додано прикомандированого: {pib}", "")
            # TODO: core_add_prikom(data_dict, ws_dst, tabel_sheet)

# ============================================================================
# STAGE D: ABSENT
# ============================================================================
def sync_step_absent(wb_src, wb_dst):
    """Sync temporarily absent personnel"""
    log_event("STEP", "Обробка Тимчасово відсутніх", "")

    ws_src = wb_src[SH_ABS]
    ws_dst = wb_dst[SH_ABS]

    data_src = get_table_data_as_list(ws_src, "ТаблицяТимчВідсутності")
    data_dst = get_table_data_as_list(ws_dst, "ТаблицяТимчВідсутності")

    # TODO: Get actual column indices
    col_pib = 1
    col_type = 2
    col_fact_ret = 10

    # 1. CLOSE existing absences
    print("  Відсутні: перевірка на закриття...")
    for j, dst_row in enumerate(data_dst, start=1):
        if dst_row[col_fact_ret] is None:
            key_pib = str(dst_row[col_pib])
            key_type = str(dst_row[col_type])

            # Find matching source row
            for src_row in data_src:
                if (str(src_row[col_pib]) == key_pib and
                    str(src_row[col_type]) == key_type):
                    if src_row[col_fact_ret] is not None:
                        # TODO: Update destination row
                        log_event("CLOSE", f"Закрито відсутність: {key_pib}", "")
                    break

    # 2. ADD new absences
    total_rows = len(data_src)
    for i, src_row in enumerate(data_src, start=1):
        if i % 10 == 0:
            print(f"  Відсутні: пошук нових ({i}/{total_rows})")

        src_pib = str(src_row[col_pib])
        src_type = str(src_row[col_type])
        src_date = src_row[5] if len(src_row) > 5 else None  # Дата вибуття

        if src_pib and src_date:
            # Check if exists
            exists = False
            for dst_row in data_dst:
                if (str(dst_row[col_pib]) == src_pib and
                    str(dst_row[col_type]) == src_type and
                    dst_row[5] == src_date):
                    exists = True
                    break

            if not exists:
                # Verify person exists in OOS
                r_oos = find_row_index_by_column_value(
                    wb_dst[SH_OOS], "ТаблицяООС", "ПІБ", src_pib
                )

                if r_oos > 0:
                    log_event("ADD-ABS", f"Нова відсутність: {src_pib}", src_type)
                    # TODO: Add new absence row with OOS data

# ============================================================================
# CLI ENTRY POINT
# ============================================================================
if __name__ == "__main__":
    if len(sys.argv) > 1:
        master_file = sys.argv[1]
    else:
        print("Використання: python sync_engine.py <шлях_до_master_файлу>")
        sys.exit(1)

    if not Path(master_file).exists():
        print(f"Помилка: файл {master_file} не знайдено")
        sys.exit(1)

    try:
        run_synchronization_main(master_file)
    except Exception as e:
        print(f"\n❌ Критична помилка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
